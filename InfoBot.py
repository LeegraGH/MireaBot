import math
import re
import matplotlib
import vk_api
from matplotlib import pyplot as plt
from vk_api import VkUpload
from vk_api.longpoll import VkLongPoll, VkEventType
import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime
import locale
import PIL.Image as Image
from PIL.Image import Resampling
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from pathlib import Path

locale.setlocale(locale.LC_ALL, "ru")


class User:
    def __init__(self, id):
        self.id = id
        self.group = ""
        self.safeGroup = ""
        self.table = ""
        self.teacher_table = ""
        self.teacher = ""
        self.message = ""
        self.last_message = ""
        self.add_message=""


nameDays = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота", "воскресенье"]
weekday = {1: "понедельник", 2: 'вторник', 3: 'среда', 4: 'четверг', 5: 'пятница', 6: 'суббота', 7: 'воскресенье'}
region_corona = {}
allSheet = []

# Время
today = datetime.date.today().strftime("%a").lower()
tomorrow = (datetime.date.today() + datetime.timedelta(days=1)).strftime("%a").lower()
thisWeek = datetime.date.today() - datetime.timedelta(days=datetime.date.today().weekday())
nextWeek = (datetime.date.today() + datetime.timedelta(weeks=1) - datetime.timedelta(
    days=datetime.date.today().weekday()))

# API-ключ
token = "58549c03604a0c74ad92e0c1eda3b02a2481ec5fb1e103222779863a1aa208540679f0c56a906f13870a9"

# Авторизация как сообщества
vk_session = vk_api.VkApi(token=token)
vk = vk_session.get_api()
upload = VkUpload(vk_session)
users = []

# Работа с сообщениями
longpoll = VkLongPoll(vk_session)

# Клавиатура
keyboard = VkKeyboard(one_time=True)
keyboard.add_button("на сегодня", color=VkKeyboardColor.POSITIVE)
keyboard.add_button("на завтра", color=VkKeyboardColor.NEGATIVE)
keyboard.add_line()
keyboard.add_button("на эту неделю", color=VkKeyboardColor.PRIMARY)
keyboard.add_button("на следующую неделю", color=VkKeyboardColor.PRIMARY)
keyboard.add_line()
keyboard.add_button("какая неделя?", color=VkKeyboardColor.SECONDARY)
keyboard.add_button("какая группа?", color=VkKeyboardColor.SECONDARY)

weatherKeyboard = VkKeyboard(one_time=True)
weatherKeyboard.add_button("сейчас", color=VkKeyboardColor.PRIMARY)
weatherKeyboard.add_button("сегодня", color=VkKeyboardColor.POSITIVE)
weatherKeyboard.add_button("завтра", color=VkKeyboardColor.POSITIVE)
weatherKeyboard.add_line()
weatherKeyboard.add_button("на 5 дней", color=VkKeyboardColor.POSITIVE)

keyboard1 = VkKeyboard(one_time=True)
keyboard1.add_button("на сегодня", color=VkKeyboardColor.POSITIVE)
keyboard1.add_button("на завтра", color=VkKeyboardColor.NEGATIVE)
keyboard1.add_line()
keyboard1.add_button("на эту неделю", color=VkKeyboardColor.PRIMARY)
keyboard1.add_button("на следующую неделю", color=VkKeyboardColor.PRIMARY)

# Какая неделя?
weekStudy = int(datetime.date.today().strftime("%V")) - int(datetime.date(2022, 2, 9).strftime("%V")) + 1


# Неделя по дате
def numWeek(d=datetime.date.today()):
    return int(d.strftime("%V")) - int(datetime.date(2022, 2, 9).strftime("%V")) + 1


# четность недели
def getParity(weekStudy):
    if weekStudy % 2 == 0:
        return 0
    else:
        return 1


def changeMonth(m):
    out = ''
    if m == 'январь':
        out = 'января'
    elif m == 'декабрь':
        out = 'декабря'
    elif m == 'февраль':
        out = 'февраля'
    elif m == 'март':
        out = 'марта'
    elif m == 'апрель':
        out = 'апреля'
    elif m == 'май':
        out = 'мая'
    elif m == 'июнь':
        out = 'июня'
    elif m == 'июль':
        out = 'июля'
    elif m == 'август':
        out = 'августа'
    elif m == 'сентябрь':
        out = 'сентябрь'
    elif m == 'октябрь':
        out = 'октября'
    elif m == 'ноябрь':
        out = 'ноября'
    elif m == 'декабрь':
        out = 'декабря'
    return out


def typeWind(wind):
    if 0 <= wind < 0.3:
        return "штиль"
    elif 0.3 <= wind < 1.6:
        return "тихий"
    elif 1.6 <= wind < 3.4:
        return "лёгкий"
    elif 3.4 <= wind < 5.5:
        return "слабый"
    elif 5.5 <= wind < 8:
        return "умеренный"
    elif 8 <= wind < 10.8:
        return "свежий"
    elif 10.8 <= wind < 13.9:
        return "сильный"
    elif 13.9 <= wind < 17.2:
        return "крепкий"
    elif 17.2 <= wind < 20.8:
        return "очень крепкий"
    elif 20.8 <= wind < 24.5:
        return "шторм"
    elif 24.5 <= wind < 28.5:
        return "сильный шторм"
    elif 28.5 <= wind < 33:
        return "жестокий шторм"
    else:
        return "ураган"


def wayWind(deg):
    if 0 <= deg < 45:
        return "северный"
    elif 45 <= deg < 90:
        return "северо-восточный"
    elif 90 <= deg < 135:
        return "восточный"
    elif 135 <= deg < 180:
        return "юго-восточный"
    elif 180 <= deg < 225:
        return "южный"
    elif 225 <= deg < 270:
        return "юго-западный"
    elif 270 <= deg < 315:
        return "западный"
    else:
        return "северо-западный"


def setCourse(group):
    if datetime.date.today().month in [9, 10, 11, 12]:
        if int(group[8:]) == datetime.date.today().year:
            return 1
        elif int(datetime.date.today().strftime("%y")) - int(group[8:]) == 1:
            return 2
        else:
            return 3
    else:
        if int(datetime.date.today().strftime("%y")) - int(group[8:]) == 1:
            return 1
        elif int(datetime.date.today().strftime("%y")) - int(group[8:]) == 2:
            return 2
        else:
            return 3


# Расписание с сайта
page = requests.get("https://www.mirea.ru/schedule/")
soup = BeautifulSoup(page.text, "html.parser")

result = soup.find('div', {'class': 'rasspisanie'}).find(string="Институт информационных технологий").find_parent(
    'div').find_parent('div').findAll(href=re.compile('https://'))[:3]

for i in range(len(result)):
    res = result[i].get('href')
    f = open("file" + str(i + 1) + ".xlsx", "wb")
    resp = requests.get(res)
    f.write(resp.content)
    book = openpyxl.load_workbook("file" + str(i + 1) + ".xlsx")
    sheet = book.active
    allSheet.append(sheet)

# Погода
weather_api1 = requests.get(
    "http://api.openweathermap.org/data/2.5/weather?q=moscow&appid=cedff4064e83e3f289965131db177a37&units=metric&lang=ru")
weather1 = weather_api1.json()
weather_api2 = requests.get(
    "http://api.openweathermap.org/data/2.5/forecast?lat=55.7522&lon=37.6156&appid=cedff4064e83e3f289965131db177a37&units=metric&lang=ru")
weather2 = weather_api2.json()


# Расписание группы
def groupTimetable(date,table,group):
    dayWeek = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб']
    num_cols = sheet.max_column
    for k in range(1, num_cols):
        if sheet.cell(row=2, column=k).value == group:
            num = 1
            if date == dayWeek[0]:
                diapason = sheet.iter_rows(min_row=4, max_row=15, min_col=k, max_col=k + 3)
            elif date == dayWeek[1]:
                diapason = sheet.iter_rows(min_row=16, max_row=27, min_col=k, max_col=k + 3)
            elif date == dayWeek[2]:
                diapason = sheet.iter_rows(min_row=28, max_row=39, min_col=k, max_col=k + 3)
            elif date == dayWeek[3]:
                diapason = sheet.iter_rows(min_row=40, max_row=51, min_col=k, max_col=k + 3)
            elif date == dayWeek[4]:
                diapason = sheet.iter_rows(min_row=52, max_row=63, min_col=k, max_col=k + 3)
            else:
                diapason = sheet.iter_rows(min_row=64, max_row=75, min_col=k, max_col=k + 3)

            parity = getParity(weekStudy)
            if parity == 0:
                m = 0
                for row in diapason:
                    if m % 2 == 0:
                        m += 1
                        continue
                    table += str(num) + ") "
                    num += 1
                    for c in range(len(row)):
                        if row[c].value is None or row[c].value.upper() == row[c].value.lower():
                            table += '-\n'
                            break
                        else:
                            if row[-1] != row[c]:
                                if "\n" in row[c].value and c == 0:
                                    for b in range(len(row[c].value.split("\n"))):
                                        table += row[c].value.split("\n")[b] + ", " + row[c + 1].value.split("\n")[
                                            b] + ", " + row[c + 2].value.split("\n")[b] + ", " + \
                                                 row[c + 3].value.split("\n")[b] + '\n'
                                    break
                                else:
                                    if row[-2] == row[c] and "\n" in row[c].value:
                                        for w in range(len(row[c].value.split("\n"))):
                                            table += row[c].value.split("\n")[w] + ", "
                                    else:
                                        table += row[c].value + ", "
                            else:
                                table += row[c].value + '\n'
                    m += 1
            elif parity == 1:
                m = 0
                for row in diapason:
                    if m % 2 != 0:
                        m += 1
                        continue
                    table += str(num) + ") "
                    num += 1
                    for c in range(len(row)):
                        if row[c].value is None or row[c].value.upper() == row[c].value.lower():
                            table += '-\n'
                            break
                        else:
                            if row[-1] != row[c]:
                                if "\n" in row[c].value and c == 0:
                                    for b in range(len(row[c].value.split("\n"))):
                                        table += row[c].value.split("\n")[b] + ", " + row[c + 1].value.split("\n")[
                                            b] + ", " + row[c + 2].value.split("\n")[b] + ", " + \
                                                 row[c + 3].value.split("\n")[b] + '\n'
                                    break
                                else:
                                    if row[-2] == row[c] and "\n" in row[c].value:
                                        for w in range(len(row[c].value.split("\n"))):
                                            table += row[c].value.split("\n")[w] + ", "
                                    else:
                                        table += row[c].value + ", "
                            else:
                                table += row[c].value + '\n'
                    m += 1
    return table


# Поиск преподавателя
def searchTeacher(teacher):
    same_teachers = []
    for sheet in allSheet:
        num_cols = sheet.max_column
        num_rows = sheet.max_row
        for j in range(1, num_cols):
            if sheet.cell(row=3, column=j).value == "ФИО преподавателя":
                for h in range(4, num_rows):
                    if sheet.cell(row=h, column=j).value is not None:
                        if '\n' in str(sheet.cell(row=h, column=j).value):
                            temp_teachers = str(sheet.cell(row=h, column=j).value).split('\n')
                        else:
                            temp_teachers = str(sheet.cell(row=h, column=j).value).split(', ')
                        for v in temp_teachers:
                            if (v[:-5] == teacher or v == teacher or v[:-2] == teacher) and v not in same_teachers:
                                same_teachers.append(v)
                            elif "." not in v and (
                                    v == teacher or v == teacher[:-5] or v == teacher[:-2]) and v not in same_teachers:
                                same_teachers.append(v)
    return same_teachers


# Расписание преподавателя
def teacherTimetable(teacher):
    global weekday
    rasp = [{}, {}]
    for i in range(2):
        for j in range(6):
            rasp[i][weekday[j + 1]] = ['-'] * 6
    for sheet in allSheet:
        num_cols = sheet.max_column
        num_rows = sheet.max_row
        for i in range(1, num_cols + 1):
            if sheet.cell(row=3, column=i).value == "ФИО преподавателя":
                for j in range(3, num_rows + 1):
                    if sheet.cell(row=j, column=i).value is not None:
                        if '\n' in str(sheet.cell(row=j, column=i).value):
                            var = str(sheet.cell(row=j, column=i).value).split('\n')
                        else:
                            var = str(sheet.cell(row=j, column=i).value).split(', ')
                        for k in range(len(var)):
                            if teacher.lower() == var[k].lower().strip().lstrip():
                                num_day = weekday[(j - 4) // 12 + 1]
                                num_w = j % 2
                                num_p = ((j - 4) % 12) // 2
                                group = str(sheet.cell(row=2, column=i - 2).value)
                                if rasp[num_w][num_day][num_p] != '-':
                                    rasp[num_w][num_day][num_p] = rasp[num_w][num_day][num_p] + ", " + group
                                else:
                                    try:
                                        rasp[num_w][num_day][num_p] = group
                                        if str(sheet.cell(row=j, column=i - 1).value).split('\n')[k] is not None:
                                            rasp[num_w][num_day][num_p] = \
                                                str(sheet.cell(row=j, column=i - 1).value).split('\n')[k] + ", " + \
                                                rasp[num_w][num_day][num_p]
                                        if str(sheet.cell(row=j, column=i + 1).value).split('\n')[k] is not None:
                                            rasp[num_w][num_day][num_p] = \
                                                str(sheet.cell(row=j, column=i + 1).value).split('\n')[k] + ", " + \
                                                rasp[num_w][num_day][num_p]
                                        rasp[num_w][num_day][num_p] = str(
                                            sheet.cell(row=j, column=i - 2).value.split('\n')[k]) + ", " + \
                                                                      rasp[num_w][num_day][num_p]
                                    except IndexError:
                                        rasp[num_w][num_day][num_p] = group
                                        if str(sheet.cell(row=j, column=i - 1).value) is not None:
                                            rasp[num_w][num_day][num_p] = str(
                                                sheet.cell(row=j, column=i - 1).value) + ", " + rasp[num_w][num_day][
                                                                              num_p]
                                        if str(sheet.cell(row=j, column=i + 1).value) is not None:
                                            rasp[num_w][num_day][num_p] = str(
                                                sheet.cell(row=j, column=i + 1).value) + ", " + rasp[num_w][num_day][
                                                                              num_p]
                                        rasp[num_w][num_day][num_p] = str(
                                            sheet.cell(row=j, column=i - 2).value) + ", " + rasp[num_w][num_day][num_p]
    return rasp


# график по коронавирусу
def corona_shedule():
    corona_site = requests.get('https://coronavirusstat.ru/country/russia/')
    soup = BeautifulSoup(corona_site.text, "html.parser")
    stat_info = soup.find('table').find('tbody')
    date = []
    active = []
    cure = []
    die = []
    for i in range(10):
        date.append(stat_info.find_all('th')[i].text)
        active.append(int(stat_info.find_all('tr')[i].find_all('td')[0].text.split()[0]))
        cure.append(active[i] + int(stat_info.find_all('tr')[i].find_all('td')[1].text.split()[0]))
        die.append(cure[i] + int(stat_info.find_all('tr')[i].find_all('td')[2].text.split()[0]))
    date.reverse()
    active.reverse()
    cure.reverse()
    die.reverse()
    matplotlib.use("TkAgg")
    fig = plt.figure()
    plt.plot(date, active, "black")
    plt.plot(date, cure, "red")
    plt.ylim([0, 20000000])
    plt.title("Россия - детальная статистика - коронавирус")
    plt.plot(date, die, label="умерло", color="orange")
    plt.fill_between(date, die, color="orange")
    plt.plot(date, cure, label="вылечено", color="yellowgreen")
    plt.fill_between(date, cure, color="yellowgreen")
    plt.plot(date, active, label="активных", color="turquoise")
    plt.fill_between(date, active, color="turquoise")
    plt.xticks(rotation=25, fontsize=7)
    plt.yticks(fontsize=7)
    plt.gca().yaxis.set_major_formatter(matplotlib.ticker.FormatStrFormatter('%d'))
    plt.legend()
    plt.grid(False)
    fig.savefig('corona.png')


# Работа бота
for event in longpoll.listen():
    if event.type == VkEventType.MESSAGE_NEW:
        if event.to_me:
            id = event.user_id
            check = False
            for user in users:
                if user.id == id:
                    check = True
                    break
            if not check:
                users.append(User(id))
            for user in users:
                if user.id == id:
                    user.message = event.text.lower()
                    user.add_message = event.text
                    if user.message == "начать":
                        vk_session.method('messages.send', {'user_id': id, 'sticker_id': 8801, 'random_id': 0})
                        vk_session.method('messages.send',
                                          {'user_id': id,
                                           'message': 'Список доступных команд:\n1) Начать - посмотреть список доступных команд\n\
        2) Бот - расписание для группы (формат группы - ИКБО-08-21)\n\
        3) Бот <название группы> - расписание для указанной группы\n\
        4) Бот <день недели> <название группы> - расписание для указанной группы на определенный день недели\n\
        5) Бот <день недели> - расписание для сохраненной группы на определенный день недели\n\
        6) Найти <фамилия преподавателя> - расписание для преподавателя\n\
        7) Погода - погода в Москве\n\
        8) Корона - статистика по коронавирусу в России\n\
        9) Корона <название региона> - статистика по коронавирусу в указанном регионе',
                                           'random_id': 0})
                    elif re.search(r'[а-яА-Я]{4}-\d{2}-\d{2}', user.message) is not None and len(user.message) == 10:
                        user.last_message = ""
                        user.group = user.message.upper()
                        user.safeGroup = user.group
                        vk_session.method('messages.send',
                                          {'user_id': id, 'message': "Я запомнил, что вы из группы " + user.group,
                                           'random_id': 0})
                        course = setCourse(user.group)
                        sheet = allSheet[course - 1]
                    elif user.message == "бот":
                        user.last_message = ""
                        if user.group != "":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Показать расписание группы " + user.group,
                                               'keyboard': keyboard.get_keyboard(), 'random_id': 0})
                        else:
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Вы не уточнили группу!", 'random_id': 0})
                    elif user.message == "какая неделя?":
                        user.last_message = ""
                        vk_session.method('messages.send',
                                          {'user_id': id, 'message': "Идет " + str(weekStudy) + " неделя",
                                           'random_id': 0})
                    elif user.message == "какая группа?":
                        user.last_message = ""
                        user.group = user.safeGroup
                        if user.group != "":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Сохраненная группа - " + user.group,
                                               'random_id': 0})
                        else:
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Вы не уточнили группу!", 'random_id': 0})
                    elif user.message == "на сегодня" and user.last_message == "":
                        if user.group == "":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Вы не уточнили группу!", 'random_id': 0})
                        elif today == "вс":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Выходной :)",
                                               'random_id': 0})
                            vk_session.method('messages.send',
                                              {'user_id': id, 'sticker_id': 8797, 'random_id': 0})
                        else:
                            user.table=groupTimetable(today,user.table,user.group)
                            if user.table == "":
                                user.group = ""
                            if user.group == "":
                                vk_session.method('messages.send',
                                                  {'user_id': id,
                                                   'message': "В расписании нет такой группы! Попробуйте ввести другую группу.",
                                                   'random_id': 0})
                            else:
                                month = datetime.date.today().strftime("%B").lower()
                                month = changeMonth(month)
                                user.table = "Расписание на " + datetime.date.today().strftime(
                                    "%d") + " " + month + "\n" + user.table
                                vk_session.method('messages.send', {'user_id': id, 'message': user.table, 'random_id': 0})
                                user.table = ""
                        user.group = user.safeGroup
                    elif user.message == "на завтра" and user.last_message == "":
                        if user.group == "":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Вы не уточнили группу!", 'random_id': 0})
                        elif tomorrow == "вс":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Выходной :)",
                                               'random_id': 0})
                            vk_session.method('messages.send',
                                              {'user_id': id, 'sticker_id': 8797, 'random_id': 0})
                        else:
                            if tomorrow == "пн":
                                weekStudy += 1
                            user.table = groupTimetable(tomorrow, user.table, user.group)
                            if user.table == "":
                                user.group = ""
                            if user.group == "":
                                vk_session.method('messages.send',
                                                  {'user_id': id,
                                                   'message': "В расписании нет такой группы! Попробуйте ввести другую группу.",
                                                   'random_id': 0})
                            else:
                                month = (datetime.date.today() + datetime.timedelta(days=1)).strftime("%B").lower()
                                month = changeMonth(month)
                                user.table = "Расписание на " + (
                                        datetime.date.today() + datetime.timedelta(days=1)).strftime(
                                    "%d") + " " + month + "\n" + user.table
                                vk_session.method('messages.send', {'user_id': id, 'message': user.table, 'random_id': 0})
                                user.table = ""
                            if tomorrow == "пн":
                                weekStudy -= 1
                        user.group = user.safeGroup
                    elif user.message == "на эту неделю" and user.last_message == "":
                        if user.group == "":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Вы не уточнили группу!", 'random_id': 0})
                        else:
                            for i in range(6):
                                if i != 0:
                                    month = thisWeek.strftime("%B").lower()
                                    month = changeMonth(month)
                                    user.table = user.table + "\n" + "Расписание на " + nameDays[i] + " " + thisWeek.strftime(
                                        "%d") + " " + month + "\n"
                                user.table = groupTimetable(thisWeek.strftime("%a").lower(),user.table,user.group)
                                if user.table == "":
                                    user.group = ""
                                if user.group == "":
                                    break
                                if i == 0:
                                    month = thisWeek.strftime("%B").lower()
                                    month = changeMonth(month)
                                    user.table = "Расписание на " + nameDays[i] + " " + thisWeek.strftime(
                                        "%d") + " " + month + "\n" + user.table
                                user.table += '\n'
                                thisWeek += datetime.timedelta(days=1)
                            if user.group == "":
                                vk_session.method('messages.send',
                                                  {'user_id': id,
                                                   'message': "В расписании нет такой группы! Попробуйте ввести другую группу.",
                                                   'random_id': 0})
                            else:
                                vk_session.method('messages.send', {'user_id': id, 'message': user.table, 'random_id': 0})
                                user.table = ""
                                thisWeek = datetime.date.today() - datetime.timedelta(
                                    days=datetime.date.today().weekday())
                        user.group = user.safeGroup
                    elif user.message == "на следующую неделю" and user.last_message == "":
                        if user.group == "":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Вы не уточнили группу!", 'random_id': 0})
                        else:
                            weekStudy += 1
                            for i in range(6):
                                if i != 0:
                                    month = nextWeek.strftime("%B").lower()
                                    month = changeMonth(month)
                                    user.table = user.table + "\n" + "Расписание на " + nameDays[i] + " " + nextWeek.strftime(
                                        "%d") + " " + month + "\n"

                                user.table = groupTimetable(nextWeek.strftime("%a").lower(),user.table,user.group)
                                if user.table == "":
                                    user.group = ""
                                if user.table == "":
                                    user.group = ""
                                if user.group == "":
                                    break
                                if i == 0:
                                    month = nextWeek.strftime("%B").lower()
                                    month = changeMonth(month)
                                    user.table = "Расписание на " + nameDays[i] + " " + nextWeek.strftime(
                                        "%d") + " " + month + "\n" + user.table
                                user.table += '\n'
                                nextWeek += datetime.timedelta(days=1)
                            if user.group == "":
                                vk_session.method('messages.send',
                                                  {'user_id': id,
                                                   'message': "В расписании нет такой группы! Попробуйте ввести другую группу.",
                                                   'random_id': 0})
                            else:
                                vk_session.method('messages.send', {'user_id': id, 'message': user.table, 'random_id': 0})
                                user.table = ""
                                nextWeek = (datetime.date.today() + datetime.timedelta(weeks=1) - datetime.timedelta(
                                    days=datetime.date.today().weekday()))
                            weekStudy -= 1
                        user.group = user.safeGroup
                    elif "бот" in user.message and user.message[4:] in nameDays:
                        user.last_message = ""
                        if user.group == "":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Вы не уточнили группу!", 'random_id': 0})
                        elif user.message[4:] == "воскресенье":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "По воскресеньям не учимся :)",
                                               'random_id': 0})
                            vk_session.method('messages.send',
                                              {'user_id': id, 'sticker_id': 8802, 'random_id': 0})
                        else:
                            listDayWeek = [datetime.date(2022, 2, 21),
                                           datetime.date(2022, 2, 22),
                                           datetime.date(2022, 2, 23),
                                           datetime.date(2022, 2, 24),
                                           datetime.date(2022, 2, 25),
                                           datetime.date(2022, 2, 26)]
                            idx = 0
                            while user.message[4:] != nameDays[idx]:
                                idx += 1
                            weekStudy = 1

                            user.table = groupTimetable(listDayWeek[idx].strftime("%a").lower(),user.table,user.group)
                            if user.table == "":
                                user.group = ""
                            if user.group == "":
                                vk_session.method('messages.send',
                                                  {'user_id': id,
                                                   'message': "В расписании нет такой группы! Попробуйте ввести другую группу.",
                                                   'random_id': 0})
                            else:
                                weekStudy = 0
                                if user.message[4:] in ["понедельник", "вторник", "четверг"]:
                                    user.table = "Расписание на нечетный " + user.message[4:] + "\n" + user.table + "\n"
                                    user.table += "\nРасписание на четный " + user.message[4:] + "\n"
                                else:
                                    user.table = "Расписание на нечетную " + user.message[4:-1] + "у\n" + user.table + "\n"
                                    user.table += "\nРасписание на четную " + user.message[4:-1] + "у\n"

                                user.table = groupTimetable((listDayWeek[idx] + datetime.timedelta(weeks=1)).strftime("%a").lower(),user.table,user.group)
                                if user.table == "":
                                    user.group = ""
                                vk_session.method('messages.send', {'user_id': id, 'message': user.table, 'random_id': 0})
                                user.table = ""
                            weekStudy = int(datetime.date.today().strftime("%V")) - int(
                                datetime.date(2022, 2, 9).strftime("%V")) + 1
                    elif "бот" in user.message and user.message[
                                              user.message.find(" ") + 1:user.message.rfind(" ")] in nameDays and re.search(
                        r'[а-яА-Я]{4}-\d{2}-\d{2}', user.message) is not None:
                        user.last_message = ""
                        user.group = user.message[user.message.rfind(" ") + 1:].upper()
                        course = setCourse(user.group)
                        sheet = allSheet[course - 1]
                        if user.message[user.message.find(" ") + 1:user.message.rfind(" ")] == "воскресенье":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "По воскресеньям не учимся :)",
                                               'random_id': 0})
                            vk_session.method('messages.send',
                                              {'user_id': id, 'sticker_id': 8802, 'random_id': 0})
                        else:
                            listDayWeek = [datetime.date(2022, 2, 21),
                                           datetime.date(2022, 2, 22),
                                           datetime.date(2022, 2, 23),
                                           datetime.date(2022, 2, 24),
                                           datetime.date(2022, 2, 25),
                                           datetime.date(2022, 2, 26)]
                            idx = 0
                            while user.message[user.message.find(" ") + 1:user.message.rfind(" ")] != nameDays[idx]:
                                idx += 1
                            weekStudy = 1

                            user.table = groupTimetable(listDayWeek[idx].strftime("%a").lower(),user.table,user.group)
                            if user.table == "":
                                user.group = ""
                            if user.group == "":
                                vk_session.method('messages.send',
                                                  {'user_id': id,
                                                   'message': "В расписании нет такой группы!",
                                                   'random_id': 0})
                            else:
                                weekStudy = 0
                                if user.message[user.message.find(" ") + 1:user.message.rfind(" ")] in ["понедельник", "вторник",
                                                                                         "четверг"]:
                                    user.table = "Расписание на нечетный " + user.message[user.message.find(" "):user.message.rfind(
                                        " ")] + "\n" + user.table + "\n"
                                    user.table += "\nРасписание на четный " + user.message[
                                                                         user.message.find(" "):user.message.rfind(" ")] + "\n"
                                else:
                                    user.table = "Расписание на нечетную " + user.message[user.message.find(" "):user.message.rfind(
                                        " ") - 1] + "у\n" + user.table + "\n"
                                    user.table += "\nРасписание на четную " + user.message[
                                                                         user.message.find(" "):user.message.rfind(
                                                                             " ") - 1] + "у\n"

                                user.table = groupTimetable((listDayWeek[idx] + datetime.timedelta(weeks=1)).strftime("%a").lower(),user.table,user.group)
                                if user.table == "":
                                    user.group = ""
                                vk_session.method('messages.send', {'user_id': id, 'message': user.table, 'random_id': 0})
                                user.table = ""
                                user.group = user.safeGroup
                            weekStudy = int(datetime.date.today().strftime("%V")) - int(
                                datetime.date(2022, 2, 9).strftime("%V")) + 1
                    elif "бот" in user.message and re.search(r'[а-яА-Я]{4}-\d{2}-\d{2}', user.message) is not None and len(
                            user.message) == 14:
                        user.last_message = ""
                        user.group = user.message[user.message.find(" ") + 1:].upper()
                        course = setCourse(user.group)
                        sheet = allSheet[course - 1]
                        vk_session.method('messages.send',
                                          {'user_id': id, 'message': "Показать расписание группы " + user.group,
                                           'keyboard': keyboard.get_keyboard(), 'random_id': 0})
                    elif user.message == "погода":
                        vk_session.method('messages.send', {'user_id': id, 'message': "Показать погоду в Москве",
                                                            'keyboard': weatherKeyboard.get_keyboard(), 'random_id': 0})
                        user.last_message = "погода"
                    elif user.last_message == "погода" and user.message == "сейчас":
                        user.last_message = ""
                        image = requests.get(
                            "http://openweathermap.org/img/wn/" + weather1['weather'][0]['icon'] + "@2x.png",
                            stream=True)
                        attachments = []
                        photo = upload.photo_messages(photos=image.raw)[0]
                        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
                        info_weather = str.capitalize(weather1['weather'][0]['description']) + ", температура: " + str(
                            round(int(weather1['main']['temp_min']))) + " - " + str(
                            round(int(weather1['main']['temp_max']))) + " °C\nДавление: " + \
                                       str(round(
                                           int(weather1['main'][
                                                   'pressure']) / 1.33322)) + " мм рт.ст., влажность: " + str(
                            weather1['main'][
                                'humidity']) + "%\nВетер: " + typeWind(float(weather1['wind']['speed'])) + ", " + \
                                       str(weather1['wind'][
                                               'speed']) + " м/c, " + wayWind(float(weather1['wind']['deg']))
                        vk.messages.send(
                            message="Погода в Москве:\n",
                            random_id=0,
                            peer_id=id,
                            attachment=','.join(attachments)
                        )
                        vk_session.method('messages.send', {'user_id': id, 'message': info_weather, 'random_id': 0})
                    elif user.last_message == "погода" and (user.message == "завтра" or user.message == "сегодня"):
                        user.last_message = ""
                        info_weather = ""
                        four_kind_day = ["УТРО", "ДЕНЬ", "ВЕЧЕР", "НОЧЬ"]
                        time_moments = ["06:00:00", "12:00:00", "18:00:00", "00:00:00"]
                        temp = "|| "
                        num_pictures = 0
                        for i in range(len(weather2["list"])):
                            if (user.message == "завтра" and weather2["list"][i]["dt_txt"] == (
                                    datetime.date.today() + datetime.timedelta(days=1)).strftime(
                                "%Y-%m-%d") + " 06:00:00") or (
                                    user.message == "сегодня" and weather2["list"][i]["dt_txt"][
                                                             :10] == datetime.date.today().strftime("%Y-%m-%d") and int(
                                weather2["list"][i]["dt_txt"][11:13]) % 6 == 0 and int(
                                weather2["list"][i]["dt_txt"][11:13]) != 0) or (
                                    user.message == "сегодня" and weather2["list"][i]["dt_txt"] == (
                                    datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y-%m-%d") + " 00:00:00"):
                                for k in range(4):
                                    if weather2["list"][i]["dt_txt"][11:] != time_moments[k]:
                                        continue
                                    if user.message == "сегодня" and weather2["list"][i]["dt_txt"][:10] == (
                                            datetime.date.today() + datetime.timedelta(days=1)).strftime("%Y-%m-%d") and \
                                            weather2["list"][i]["dt_txt"][11:] != "00:00:00":
                                        break
                                    info_weather += four_kind_day[k] + "\n" + str.capitalize(
                                        weather2["list"][i]['weather'][0]['description']) + ", температура: " + str(
                                        math.floor(float(weather2["list"][i]['main']['temp_min']))) + " - " + str(
                                        math.ceil(float(weather2["list"][i]['main']['temp_max']))) + " °C\nДавление: " + \
                                                    str(round(int(
                                                        weather2["list"][i]['main'][
                                                            'pressure']) / 1.33322)) + " мм рт.ст., влажность: " + str(
                                        weather2["list"][i]['main'][
                                            'humidity']) + "%\nВетер: " + typeWind(
                                        float(weather2["list"][i]['wind']['speed'])) + ", " + \
                                                    str(weather2["list"][i]['wind'][
                                                            'speed']) + " м/c, " + wayWind(
                                        float(weather2["list"][i]['wind']['deg']))
                                    info_weather += "\n\n"
                                    temp += str(round(float(weather2["list"][i]['main']['temp']))) + "°C || "
                                    image = requests.get(
                                        "http://openweathermap.org/img/wn/" + weather2["list"][i]['weather'][0][
                                            'icon'] + "@2x.png", stream=True)
                                    if user.message == "сегодня":
                                        with open("file" + str(k + 1) + "+.png", "wb") as f:
                                            f.write(image.content)
                                        num_pictures += 1
                                    else:
                                        with open("file" + str(k + 1) + ".png", "wb") as f:
                                            f.write(image.content)
                                    i += 2
                                break
                        if user.message == "сегодня":
                            img = Image.new('RGBA', ((int(str(num_pictures) + "00")), 100))
                        else:
                            img = Image.new('RGBA', (400, 100))
                        if user.message == "сегодня":
                            h = -1
                            for t in range(4):
                                nameFile = "file" + str(t + 1) + "+.png"
                                if not Path(nameFile).is_file():
                                    continue
                                h += 1
                                imgN = Image.open(nameFile)
                                img.paste(imgN, (0 + 100 * h, 0), imgN)
                        else:
                            for t in range(4):
                                nameFile = "file" + str(t + 1) + ".png"
                                imgN = Image.open(nameFile)
                                img.paste(imgN, (0 + 100 * t, 0), imgN)
                        if user.message == "завтра":
                            img = img.resize((180, 40), Resampling.LANCZOS)
                        else:
                            if num_pictures == 1:
                                img = img.resize((50, 50), Resampling.LANCZOS)
                            elif num_pictures == 2:
                                img = img.resize((80, 40), Resampling.LANCZOS)
                            elif num_pictures == 3:
                                img = img.resize((130, 40), Resampling.LANCZOS)
                            else:
                                img = img.resize((180, 40), Resampling.LANCZOS)
                        img.save("image.png")
                        attachments = []
                        photo = upload.photo_messages(photos="image.png")[0]
                        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
                        if user.message == "завтра":
                            vk.messages.send(
                                message="Погода в Москве завтра:\n",
                                random_id=0,
                                peer_id=id,
                                attachment=','.join(attachments)
                            )
                        else:
                            vk.messages.send(
                                message="Погода в Москве сегодня:\n",
                                random_id=0,
                                peer_id=id,
                                attachment=','.join(attachments)
                            )
                        vk_session.method('messages.send', {'peer_id': id, 'message': temp, 'random_id': 0})
                        vk_session.method('messages.send', {'peer_id': id, 'message': info_weather, 'random_id': 0})
                    elif user.last_message == "погода" and user.message == "на 5 дней":
                        user.last_message = ""
                        day_weather = "| "
                        night_weather = "| "
                        img = Image.new('RGBA', (500, 100))
                        t = 0
                        m = 0
                        flag = False
                        for i in range(len(weather2["list"])):
                            if weather2["list"][i]["dt_txt"][:10] == datetime.date.today().strftime("%Y-%m-%d") and int(
                                    weather2["list"][i]["dt_txt"][11:13]) > 15 and t < 5 and not flag:
                                flag = True
                                day_weather += "✖" + " | "
                                t += 1
                            elif weather2["list"][i]["dt_txt"][11:] == "15:00:00" and t < 5:
                                flag = True
                                temp_day = str(round(weather2["list"][i]["main"]["temp"])) if len(
                                    str(round(weather2["list"][i]["main"]["temp"]))) == 2 else "+" + str(
                                    round(weather2["list"][i]["main"]["temp"]))
                                day_weather += temp_day + "°C | "
                                t += 1
                                image = requests.get(
                                    "http://openweathermap.org/img/wn/" + weather2["list"][i]['weather'][0][
                                        'icon'] + "@2x.png",
                                    stream=True)
                                with open("file" + str(t) + "5.png", "wb") as f:
                                    f.write(image.content)
                            elif weather2["list"][i]["dt_txt"][11:] == "00:00:00" and weather2["list"][i]["dt_txt"][
                                                                                      :10] != datetime.date.today().strftime(
                                "%Y-%m-%d"):
                                temp_night = str(round(weather2["list"][i]["main"]["temp"])) if len(
                                    str(round(weather2["list"][i]["main"]["temp"]))) == 2 else "+" + str(
                                    round(weather2["list"][i]["main"]["temp"]))
                                night_weather += temp_night + "°C | "
                                m += 1
                        if m == 4:
                            night_weather += "✖ | "
                        for u in range(5):
                            nameFile = "file" + str(u + 1) + "5.png"
                            if not Path(nameFile).is_file():
                                continue
                            imgN = Image.open(nameFile)
                            img.paste(imgN, (0 + 100 * u, 0), imgN)
                        img = img.resize((200, 40), Resampling.LANCZOS)
                        img.save("image.png")
                        attachments = []
                        photo = upload.photo_messages(photos="image.png")[0]
                        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
                        vk.messages.send(
                            message="Погода в Москве с " + datetime.date.today().strftime(
                                "%d.%m") + "-" + (datetime.date.today() + datetime.timedelta(days=4)).strftime("%d.%m"),
                            random_id=0,
                            peer_id=id,
                            attachment=','.join(attachments)
                        )
                        vk_session.method('messages.send',
                                          {'peer_id': id, 'message': day_weather + "ДЕНЬ", 'random_id': 0})
                        vk_session.method('messages.send',
                                          {'peer_id': id, 'message': night_weather + "НОЧЬ", 'random_id': 0})
                    elif "найти" in user.message and len(user.message) > 6:
                        user.last_message = ""
                        if "." not in user.message:
                            user.teacher = str.capitalize(user.message[6:])
                        else:
                            user.teacher = str.capitalize(user.message[6:user.message.find(".") - 2]) + str.upper(
                                user.message[user.message.find(".") - 2:])
                        same_teachers = searchTeacher(user.teacher)
                        if len(same_teachers) == 0:
                            vk_session.method('messages.send',
                                              {'peer_id': id, 'message': "Такого преподавателя нет!", 'random_id': 0})
                            user.teacher = ""
                        elif len(same_teachers) == 1:
                            user.teacher = same_teachers[0]
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Показать расписание преподавателя " + user.teacher,
                                               'keyboard': keyboard1.get_keyboard(), 'random_id': 0})
                            user.last_message = "Показать расписание преподавателя"
                        else:
                            nameKeyboard = VkKeyboard(one_time=True)
                            for name in range(len(same_teachers)):
                                nameKeyboard.add_button(label=same_teachers[name], color=VkKeyboardColor.PRIMARY)
                                if name != len(same_teachers) - 1:
                                    nameKeyboard.add_line()
                            vk_session.method('messages.send', {'user_id': id, 'message': "Выберите преподавателя",
                                                                'keyboard': nameKeyboard.get_keyboard(),
                                                                'random_id': 0})
                            user.last_message = "Выберите преподавателя"
                    elif user.last_message == "Выберите преподавателя":
                        user.teacher = user.add_message
                        vk_session.method('messages.send',
                                          {'user_id': id, 'message': "Показать расписание преподавателя " + user.teacher,
                                           'keyboard': keyboard1.get_keyboard(), 'random_id': 0})
                        user.last_message = "Показать расписание преподавателя"
                    elif user.last_message == "Показать расписание преподавателя" and user.message == "на сегодня":
                        if today == "вс":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Выходной :)",
                                               'random_id': 0})
                            vk_session.method('messages.send',
                                              {'user_id': id, 'sticker_id': 8797, 'random_id': 0})
                        else:
                            ras = teacherTimetable(user.teacher)
                            month = datetime.date.today().strftime("%B").lower()
                            month = changeMonth(month)
                            d = datetime.date.today()
                            user.teacher_table += "Расписание преподавателя " + user.teacher + " " + d.strftime(
                                '%d ') + month + '\n'
                            for k in range(6):
                                user.teacher_table += str(k + 1) + ") " + \
                                                 ras[(numWeek(d) + 1) % 2][weekday[int(d.isoweekday())]][k] + '\n'
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': user.teacher_table, 'random_id': 0})
                        user.teacher_table = ""
                        user.last_message = ""
                    elif user.last_message == "Показать расписание преподавателя" and user.message == "на завтра":
                        if tomorrow == "вс":
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': "Выходной :)",
                                               'random_id': 0})
                            vk_session.method('messages.send',
                                              {'user_id': id, 'sticker_id': 8797, 'random_id': 0})
                        else:
                            ras = teacherTimetable(user.teacher)
                            month = (datetime.date.today() + datetime.timedelta(days=1)).strftime("%B").lower()
                            month = changeMonth(month)
                            d = datetime.date.today() + datetime.timedelta(days=1)
                            user.teacher_table += "Расписание преподавателя " + user.teacher + " " + d.strftime(
                                '%d ') + month + '\n'
                            for k in range(6):
                                user.teacher_table += str(k + 1) + ") " + \
                                                 ras[(numWeek(d) + 1) % 2][weekday[int(d.isoweekday())]][
                                                     k] + '\n'
                            vk_session.method('messages.send',
                                              {'user_id': id, 'message': user.teacher_table, 'random_id': 0})
                        user.teacher_table = ""
                        user.last_message = ""
                    elif user.last_message == "Показать расписание преподавателя" and user.message == "на эту неделю":
                        ras = teacherTimetable(user.teacher)
                        user.teacher_table += "Расписание на эту неделю для преподавателя " + user.teacher + '\n'
                        d = datetime.date.today()
                        month = thisWeek.strftime("%B").lower()
                        month = changeMonth(month)
                        for i in range(1, 7):
                            user.teacher_table += "\n" + "Расписание на " + weekday[i] + " " + (
                                    thisWeek + datetime.timedelta(days=i - 1)).strftime("%d") + " " + month + "\n"
                            for j in range(6):
                                user.teacher_table += str(j + 1) + ") " + ras[(numWeek(d) + 1) % 2][weekday[i]][j] + '\n'
                        vk_session.method('messages.send', {'user_id': id, 'message': user.teacher_table, 'random_id': 0})
                        user.teacher_table = ""
                        user.last_message = ""
                    elif user.last_message == "Показать расписание преподавателя" and user.message == "на следующую неделю":
                        ras = teacherTimetable(user.teacher)
                        user.teacher_table += "Расписание на следующую неделю для преподавателя " + user.teacher + '\n'
                        d = datetime.date.today() + datetime.timedelta(weeks=1)
                        month = nextWeek.strftime("%B").lower()
                        month = changeMonth(month)
                        for i in range(1, 7):
                            user.teacher_table += "\n" + "Расписание на " + weekday[i] + " " + (
                                    nextWeek + datetime.timedelta(days=i - 1)).strftime("%d") + " " + month + "\n"
                            for j in range(6):
                                user.teacher_table += str(j + 1) + ") " + ras[(numWeek(d) + 1) % 2][weekday[i]][j] + '\n'
                        vk_session.method('messages.send', {'user_id': id, 'message': user.teacher_table, 'random_id': 0})
                        user.teacher_table = ""
                        user.last_message = ""
                    elif user.message == "корона":
                        try:
                            corona_site = requests.get("https://coronavirusstat.ru/country/russia/")
                            corona_soup = BeautifulSoup(corona_site.text, "html.parser")
                            list_corona = corona_soup.findAll('div', {'class': 'col-12 p-0'})
                            for i in list_corona:
                                reg = i.find('a').text
                                site = i.find('a').get('href')
                                region_corona[reg] = site
                            corona_date = corona_soup.find('div', {'class': 'border rounded mt-3 mb-3 p-3'}).find('h6',
                                                                                                                  {
                                                                                                                      'class': 'text-muted'}).text[
                                          :-17]
                            corona_result = corona_date + "\nСлучаев: " + corona_soup.find('div',
                                                                                           {
                                                                                               'title': 'Короновирус Россия: Случаев'}).text + " (" + corona_soup.find(
                                'span', {
                                    'class': 'font-weight-bold text-text-dark'}).text + " за сегодня)\n" + "Активных: " + corona_soup.find(
                                'div', {'title': 'Короновирус Россия: Активных'}).text + " (" + corona_soup.find('span',
                                                                                                                 {
                                                                                                                     'class': 'font-weight-bold text-primary'}).text + " за сегодня)\n" + "Вылечено: " + corona_soup.find(
                                'div', {'title': 'Короновирус Россия: Вылечено'}).text + " (" + corona_soup.find(
                                'span', {
                                    'class': 'font-weight-bold text-success'}).text + " за сегодня)\n" + "Умерло: " + corona_soup.find(
                                'div', {'title': 'Короновирус Россия: Умерло'}).text + " (" + corona_soup.find('span', {
                                'class': 'font-weight-bold text-danger'}).text + " за сегодня)"
                            vk_session.method('messages.send',
                                              {'peer_id': id, 'message': corona_result, 'random_id': 0})
                            corona_shedule()
                            attachments = []
                            photo = upload.photo_messages(photos="corona.png")[0]
                            attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
                            vk.messages.send(
                                random_id=0,
                                peer_id=id,
                                attachment=','.join(attachments)
                            )
                        except ConnectionError or AttributeError:
                            vk_session.method('messages.send',
                                              {'peer_id': id, 'message': "Нет соединения с сервером...",
                                               'random_id': 0})
                            vk_session.method('messages.send', {'peer_id': id, 'sticker_id': 11748, 'random_id': 0})
                    elif "корона" in user.message and len(user.message) >= 10:
                        answer_site = True
                        try:
                            corona_site = requests.get("https://coronavirusstat.ru/country/russia/")
                            corona_soup = BeautifulSoup(corona_site.text, "html.parser")
                            list_corona = corona_soup.findAll('div', {'class': 'col-12 p-0'})
                            for i in list_corona:
                                reg = i.find('a').text
                                site = i.find('a').get('href')
                                region_corona[reg] = site
                        except ConnectionError or AttributeError:
                            answer_site = False
                        if answer_site:
                            user_reg = user.message[user.message.find(" ") + 1:]
                            user_site = ''
                            for r in region_corona:
                                if user_reg.lower() in r.lower():
                                    user_reg = r
                                    user_site = region_corona.get(r)
                                    break
                            if user_site == '':
                                vk_session.method('messages.send',
                                                  {'peer_id': id, 'message': "Регион не найден!", 'random_id': 0})
                            else:
                                try:
                                    reg_site = requests.get("https://coronavirusstat.ru" + user_site)
                                    reg_soup = BeautifulSoup(reg_site.text, "html.parser")
                                    reg_date = reg_soup.find('div', {'class': 'border rounded mt-3 mb-3 p-3'}).find(
                                        'h6', {
                                            'class': 'text-muted'}).text[:-17]
                                    reg_name = "Короновирус " + reg_soup.find('div',
                                                                              {
                                                                                  'class': 'border rounded mt-3 mb-3 p-3'}).find(
                                        'h1',
                                        {'class': 'h2 font-weight-bold'}).text[12:]
                                    reg_result = reg_date + "\nРегион: " + user_reg + "\nСлучаев: " + reg_soup.find(
                                        'div', {
                                            'title': reg_name + ": Случаев"}).text + " (" + reg_soup.find(
                                        'span', {
                                            'class': 'font-weight-bold text-text-dark'}).text + " за сегодня)\n" + "Активных: " + reg_soup.find(
                                        'div', {'title': reg_name + ": Активных"}).text + " (" + reg_soup.find('span', {
                                        'class': 'font-weight-bold text-primary'}).text + " за сегодня)\n" + "Вылечено: " + reg_soup.find(
                                        'div', {'title': reg_name + ": Вылечено"}).text + " (" + reg_soup.find(
                                        'span', {
                                            'class': 'font-weight-bold text-success'}).text + " за сегодня)\n" + "Умерло: " + reg_soup.find(
                                        'div', {'title': reg_name + ": Умерло"}).text + " (" + reg_soup.find('span', {
                                        'class': 'font-weight-bold text-danger'}).text + " за сегодня)"
                                    vk_session.method('messages.send',
                                                      {'peer_id': id, 'message': reg_result, 'random_id': 0})
                                except ConnectionError or AttributeError:
                                    vk_session.method('messages.send',
                                                      {'peer_id': id, 'message': "Нет соединения с сервером...",
                                                       'random_id': 0})
                                    vk_session.method('messages.send',
                                                      {'peer_id': id, 'sticker_id': 11748, 'random_id': 0})
                        else:
                            vk_session.method('messages.send',
                                              {'peer_id': id, 'message': "Нет соединения с сервером...",
                                               'random_id': 0})
                            vk_session.method('messages.send', {'peer_id': id, 'sticker_id': 11748, 'random_id': 0})
                    else:
                        vk_session.method('messages.send', {'peer_id': id, 'sticker_id': 13897, 'random_id': 0})
